# backend/schema_parser.py
# Parses the _Schema sheet or modern Input/Output sheets from an Excel file into a config dict

import openpyxl
from pathlib import Path
import re


def parse_schema(xlsx_path: Path) -> dict:
    """
    Router that inspects the workbook and delegates to the appropriate parser.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheetnames = wb.sheetnames
    wb.close()

    if "_Schema" in sheetnames:
        return _parse_legacy_schema(xlsx_path)
    elif "Input" in sheetnames and "Output" in sheetnames:
        return _parse_io_schema(xlsx_path)
    else:
        raise ValueError(
            "Unsupported workbook format. The file must contain either a '_Schema' sheet "
            "(legacy format) or both 'Input' and 'Output' sheets (modern format)."
        )


def _parse_io_schema(xlsx_path: Path) -> dict:
    """
    Modern parser that maps Input and Output sheets to the config dict.
    Groups inputs relationally by their Coverage Name and hides static lookups.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws_in = wb["Input"]
    ws_out = wb["Output"]

    inputs = []
    outputs = []

    in_headers = {cell.value: idx for idx, cell in enumerate(ws_in[1]) if cell.value}
    out_headers = {cell.value: idx for idx, cell in enumerate(ws_out[1]) if cell.value}

    # 1. PRE-COMPUTATION PASS: Map UnitId to Human-Readable Names
    unit_names = {}
    for row in ws_in.iter_rows(min_row=2):
        unit_id = row[in_headers.get("UnitId")].value if "UnitId" in in_headers else None
        param_code = row[in_headers.get("ParameterCode")].value if "ParameterCode" in in_headers else None
        param_value = row[in_headers.get("ParameterValue")].value if "ParameterValue" in in_headers else None
        
        if unit_id and param_code == "COV_NAME" and param_value is not None:
            unit_names[str(unit_id)] = str(param_value).strip()

    # 2. MAIN INPUT PARSING
    for row in ws_in.iter_rows(min_row=2):
        input_key = row[in_headers.get("InputKey")].value if "InputKey" in in_headers else None
        edit_target = row[in_headers.get("EditTarget")].value if "EditTarget" in in_headers else None
        
        if not input_key or not edit_target:
            continue

        param_code = row[in_headers.get("ParameterCode")].value if "ParameterCode" in in_headers else None
        
        # Filtering A: Skip metadata rows so they don't render as editable text boxes
        if param_code in ["COV_CODE", "COV_NAME"]:
            continue

        param_value = row[in_headers.get("ParameterValue")].value if "ParameterValue" in in_headers else None
        
        # Filtering B: Skip standard formulas and ArrayFormulas in the Input sheet
        if type(param_value).__name__ == "ArrayFormula" or (isinstance(param_value, str) and param_value.startswith("=")):
            continue
            
        # Filtering C: Exclude Static Lookups by inspecting the OriginalSource
        original_source = row[in_headers.get("OriginalSource")].value if "OriginalSource" in in_headers else None
        if original_source and isinstance(original_source, str) and "!" in original_source:
            try:
                source_sheet, source_cell = original_source.split("!", 1)
                source_sheet = source_sheet.strip("'\"")
                if source_sheet in wb.sheetnames:
                    actual_val = wb[source_sheet][source_cell].value
                    # If the original cell is a formula referencing another sheet (like =Rates!B1), skip it
                    if isinstance(actual_val, str) and str(actual_val).startswith("=") and "!" in actual_val:
                        continue
            except Exception:
                pass

        # JSON serialization safety net
        if param_value is not None and not isinstance(param_value, (int, float, str, bool)):
            param_value = str(param_value)

        data_type = row[in_headers.get("DataType")].value if "DataType" in in_headers else None
        param_name = row[in_headers.get("ParameterName")].value if "ParameterName" in in_headers else None
        unit_id = row[in_headers.get("UnitId")].value if "UnitId" in in_headers else None
        category = row[in_headers.get("Category")].value if "Category" in in_headers else None

        type_map = {
            "String": "text",
            "Number": "number",
            "Currency": "number",
            "Percentage": "number",
            "Date": "date"
        }
        mapped_type = type_map.get(str(data_type).strip(), "text") if data_type else "text"
        
        # DYNAMIC GROUPING: Assign to Coverage Name, fallback to Category (e.g. "Policy")
        if unit_id and str(unit_id) in unit_names:
            group_name = unit_names[str(unit_id)]
        elif category:
            group_name = str(category).strip()
        else:
            group_name = "General"

        label_name = str(param_name).strip() if param_name else str(input_key).strip()

        entry = {
            "field": str(input_key).strip(),
            "cell": str(edit_target).strip(),
            "type": mapped_type,
            "label": label_name,
            "group": group_name,
        }
        
        if param_value is not None:
            entry["default"] = param_value
            
        inputs.append(entry)

    # 3. OUTPUT PARSING
    for row in ws_out.iter_rows(min_row=2):
        output_key = row[out_headers.get("OutputKey")].value if "OutputKey" in out_headers else None
        fetch_cell = row[out_headers.get("FetchCell")].value if "FetchCell" in out_headers else None
        
        if not output_key or not fetch_cell:
            continue

        cov_name = row[out_headers.get("CoverageName")].value if "CoverageName" in out_headers else None
        out_key_str = str(output_key).strip()

        entry = {
            "field": out_key_str,
            "cell": str(fetch_cell).strip(),
            "type": "number",
            "label": str(cov_name).strip() if cov_name else out_key_str,
            "group": "Results"
        }
        
        if "FINAL_PREMIUM" in out_key_str.upper() and "POL-1" in out_key_str.upper():
            entry["primary"] = True
            
        outputs.append(entry)

    wb.close()

    if outputs and not any(o.get("primary") for o in outputs):
        outputs[-1]["primary"] = True

    config = {
        "sheet": "Input",
        "inputs": inputs,
        "outputs": outputs,
    }

    _inject_schedule_mode(config)
    return config


def _parse_legacy_schema(xlsx_path: Path) -> dict:
    """
    Legacy parser for workbooks utilizing the _Schema sheet format.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["_Schema"]

    rater_sheet = None
    for name in wb.sheetnames:
        if name != "_Schema":
            rater_sheet = name
            break

    inputs = []
    outputs = []
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True): 
        field = row[0] if len(row) > 0 else None
        cell = row[1] if len(row) > 1 else None
        ftype = row[2] if len(row) > 2 else None
        label = row[3] if len(row) > 3 else None
        direction = row[4] if len(row) > 4 else None
        group = row[5] if len(row) > 5 else None
        options_raw = row[6] if len(row) > 6 else None
        default_raw = row[7] if len(row) > 7 else None

        if not field or not cell:
            continue

        row_count += 1
        field = str(field).strip()
        cell = str(cell).strip()
        ftype = str(ftype).strip() if ftype else "text"
        label = str(label).strip() if label else field
        direction = str(direction).strip().lower() if direction else "input"
        group = str(group).strip() if group else "General"

        options = []
        if options_raw and str(options_raw).strip():
            raw_parts = str(options_raw).split(";")
            for part in raw_parts:
                part = part.strip()
                if not part:
                    continue
                try:
                    options.append(int(part))
                except ValueError:
                    try:
                        options.append(float(part))
                    except ValueError:
                        options.append(part)

        default = None
        if default_raw is not None and str(default_raw).strip():
            default = str(default_raw).strip()
            if ftype == "number":
                try:
                    default = float(default) if "." in default else int(default)
                except ValueError:
                    pass

        entry = {
            "field": field,
            "cell": cell,
            "type": ftype,
            "label": label,
            "group": group,
        }

        if direction == "output":
            entry["primary"] = (row_count == 1 and direction == "output") or field == "premium"
            outputs.append(entry)
        else:
            if options:
                entry["options"] = options
                entry["type"] = "dropdown"
            if default is not None:
                entry["default"] = default
            inputs.append(entry)

    wb.close()

    if row_count == 0:
        raise ValueError("_Schema sheet is empty — no field definitions found.")

    if outputs and not any(o.get("primary") for o in outputs):
        outputs[0]["primary"] = True

    config = {
        "sheet": rater_sheet or "Sheet1",
        "inputs": inputs,
        "outputs": outputs,
    }

    _inject_schedule_mode(config)
    return config


def _inject_schedule_mode(config: dict) -> None:
    """
    Heuristically infer schedule/repeater blocks from parsed inputs.
    """
    inputs = config.get("inputs") or []
    if not inputs:
        return

    row_map = {}
    for inp in inputs:
        cell = str(inp.get("cell") or "").strip().upper()
        # Ensure we only strip the sheet reference if it exists before applying the regex
        cell_ref = cell.split("!")[-1] if "!" in cell else cell
        
        m = re.match(r"^([A-Z]+)(\d+)$", cell_ref)
        if not m:
            continue
        col, row_txt = m.groups()
        row = int(row_txt)
        row_map.setdefault(row, {})[col] = inp

    schedule_rows = sorted(r for r, cols in row_map.items() if "D" in cols)
    if not schedule_rows:
        return

    blocks = []
    start = schedule_rows[0]
    prev = schedule_rows[0]
    for r in schedule_rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        blocks.append((start, prev))
        start = r
        prev = r
    blocks.append((start, prev))

    schedules = []
    for idx, (row_start, row_end) in enumerate(blocks, 1):
        length = row_end - row_start + 1
        if length < 3:
            continue

        d_meta = row_map[row_start].get("D", {})

        e_meta = {}
        for rr in range(row_start, row_end + 1):
            if "E" in row_map.get(rr, {}):
                e_meta = row_map[rr].get("E", {})
                break

        group_name = str(d_meta.get("group") or "Coverage")
        key_base = re.sub(r"[^a-z0-9]+", "_", group_name.lower()).strip("_")
        key = f"{key_base}_{row_start}_{row_end}" if key_base else f"schedule_{idx}"

        schedules.append(
            {
                "key": key,
                "title": f"{group_name} ({row_start}-{row_end})",
                "rowStart": row_start,
                "rowEnd": row_end,
                "allowBlankRows": True,
                "minActiveRows": 1,
                "columns": [
                    {
                        "field": "location",
                        "column": "D",
                        "type": d_meta.get("type", "text"),
                        "label": str(d_meta.get("label") or "Location #"),
                    },
                    {
                        "field": "expiring_risk_limit",
                        "column": "E",
                        "type": e_meta.get("type", "number"),
                        "label": str(e_meta.get("label") or "Risk Limit"),
                    },
                ],
            }
        )

    if not schedules:
        return

    config["mode"] = "schedule"
    config["writeRules"] = {
        "clearUnusedRows": True,
        "emptyCellWrite": "blank",
        "rowActivePolicy": "any-non-empty-cell",
    }
    config["schedules"] = schedules